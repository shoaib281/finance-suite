import os
import yfinance as yf
import numpy as np
from openpyxl import load_workbook
import time
import logging

from loggingInit import log_function
from utils import FOLDERPATH

logger = logging.getLogger(__name__)


@log_function
def update_stock_prices(sheetname):
    wb = load_workbook(filename)
    sheet = wb[sheetname]
    sheet["A1"] = time.time()

    for row in sheet.iter_rows(min_row=3):
        ticker_object = row[1]
        ticker = ticker_object.value

        if ticker != None:
            if ticker[0] != "-":
                stock = yf.Ticker(ticker)
                data = stock.history(period="5d")
                price = round(data["Close"][-1], 2)

                cell_name = "C" + str(ticker_object.row)

                sheet[cell_name].value = price

    wb.save(filename)


def percent_change_close_dataframe(df):
    first = df.iloc[0]["Close"]
    last = df.iloc[-1]["Close"]

    if np.isnan(first) and len(df) > 2:
        first = df.iloc[1]["Close"]

    return (last / first) - 1


@log_function
def update_price_movements(sheetname):
    wb = load_workbook(filename)
    sheet = wb[sheetname]

    for row in sheet.iter_rows(min_row=3):
        ticker_object = row[1]
        ticker_row = ticker_object.row
        ticker = ticker_object.value

        if ticker is not None:
            if ticker[0] != "-":
                stock = yf.Ticker(ticker)

                periods = ["2d", "5d", "1mo", "3mo", "ytd", "1y"]

                for i in range(len(periods)):
                    print(ticker, periods[i])
                    change = percent_change_close_dataframe(stock.history(period=periods[i]))

                    sheet.cell(row=ticker_row, column=i + 3, value=change)

    sheet["A1"] = time.time()
    wb.save(filename)


filename = FOLDERPATH / "market.xlsx"
lock_filename = FOLDERPATH / ".~lock.market.xlsx#"
sheets_fundemental = "Fundemental"
sheets_movements = "Movements"


def main():
    if not os.path.exists(lock_filename):
        try:
            update_stock_prices(sheets_fundemental)
            update_price_movements(sheets_movements)
        except Exception as e:
            logger.exception(str(e))
    else:
        logger.info("File already open")


if __name__ == "__main__":
    main()
